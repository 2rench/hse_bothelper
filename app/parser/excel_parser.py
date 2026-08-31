from pathlib import Path
import json
import re

import openpyxl
import xlrd

from app.database.group_repository import (
    save_group,
)

from app.parser.lesson_parser import (
    parse_lesson_text,
)


IGNORE_COLUMNS = {
    "дни",
    "пары",
}


def normalize_whitespace(
    text: str,
) -> str:

    text = re.sub(
        r"\n+",
        "\n",
        text,
    )

    text = re.sub(
        r"[ \t]+",
        " ",
        text,
    )

    return text.strip()


def parse_day(
    raw_day: str,
) -> tuple[str, str]:

    parts = normalize_whitespace(
        raw_day
    ).split("\n")

    if len(parts) >= 2:

        return (
            parts[0],
            parts[1],
        )

    return (
        raw_day,
        "",
    )


def parse_time(
    raw_time: str,
) -> tuple[str, str]:

    cleaned = normalize_whitespace(
        raw_time
    )

    parts = cleaned.split("\n")

    if len(parts) >= 2:

        return (
            parts[0],
            parts[1],
        )

    return (
        "",
        cleaned,
    )


# ============================================================
# Excel abstraction
# ============================================================

def _is_xlrd_sheet(
    sheet,
) -> bool:

    return hasattr(
        sheet,
        "cell_value",
    )


def _get_sheet_name(
    sheet,
) -> str:

    if _is_xlrd_sheet(sheet):

        return sheet.name

    return sheet.title


def _get_merged_cells(
    sheet,
) -> list[tuple[int, int, int, int]]:

    if _is_xlrd_sheet(sheet):

        return list(
            sheet.merged_cells
        )

    merged = []

    for range_obj in sheet.merged_cells.ranges:

        merged.append(
            (
                range_obj.min_row - 1,
                range_obj.max_row,
                range_obj.min_col - 1,
                range_obj.max_col,
            )
        )

    return merged


def get_merged_region(
    sheet,
    row_index: int,
    col_index: int,
):

    for merged in _get_merged_cells(
        sheet
    ):

        (
            row_start,
            row_end,
            col_start,
            col_end,
        ) = merged

        if (
            row_start
            <= row_index
            < row_end
            and
            col_start
            <= col_index
            < col_end
        ):

            return merged

    return None


def get_cell_value(
    sheet,
    row_index: int,
    col_index: int,
):

    if _is_xlrd_sheet(sheet):

        value = sheet.cell_value(
            row_index,
            col_index,
        )

        if value not in (
            None,
            "",
        ):

            return value

        merged = get_merged_region(
            sheet,
            row_index,
            col_index,
        )

        if not merged:

            return ""

        (
            row_start,
            _,
            col_start,
            _,
        ) = merged

        return sheet.cell_value(
            row_start,
            col_start,
        )

    cell = sheet.cell(
        row_index + 1,
        col_index + 1,
    )

    value = cell.value

    if value not in (
        None,
        "",
    ):

        return value

    merged = get_merged_region(
        sheet,
        row_index,
        col_index,
    )

    if not merged:

        return ""

    (
        row_start,
        _,
        col_start,
        _,
    ) = merged

    return sheet.cell(
        row_start + 1,
        col_start + 1,
    ).value


def is_underlined(
    sheet,
    row_index: int,
    col_index: int,
) -> bool:

    # --------------------------------------------
    # XLS
    # --------------------------------------------

    if _is_xlrd_sheet(sheet):

        cell = sheet.cell(
            row_index,
            col_index,
        )

        if cell.xf_index is None:

            return False

        try:

            xf = sheet.book.xf_list[
                cell.xf_index
            ]

            font = sheet.book.font_list[
                xf.font_index
            ]

            return font.underlined != 0

        except Exception:

            return False

    # --------------------------------------------
    # XLSX
    # --------------------------------------------

    cell = sheet.cell(
        row_index + 1,
        col_index + 1,
    )

    underline = cell.font.underline

    if not underline:

        return False

    return True


def _open_workbook(
    file_path: str,
):

    path = Path(
        file_path
    )

    if path.suffix.lower() == ".xlsx":

        return openpyxl.load_workbook(
            file_path,
            data_only=True,
        )

    return xlrd.open_workbook(
        file_path,
        formatting_info=True,
    )


def _get_sheets(
    workbook,
):

    if hasattr(
        workbook,
        "sheets",
    ):

        return workbook.sheets()

    return workbook.worksheets


def _get_nrows(
    sheet,
) -> int:

    if _is_xlrd_sheet(sheet):

        return sheet.nrows

    return sheet.max_row


def _get_ncols(
    sheet,
) -> int:

    if _is_xlrd_sheet(sheet):

        return sheet.ncols

    return sheet.max_column


# ============================================================
# Groups
# ============================================================

def _get_groups(
    sheet,
) -> dict[int, str]:

    groups = {}

    GROUPS_ROW_INDEX = 2

    for col_index in range(
        _get_ncols(sheet)
    ):

        value = get_cell_value(
            sheet,
            GROUPS_ROW_INDEX,
            col_index,
        )

        if value in (
            None,
            "",
        ):

            continue

        value = normalize_whitespace(
            str(value)
        )

        if not value:

            continue

        if value.lower() in IGNORE_COLUMNS:

            continue

        groups[
            col_index
        ] = value

        save_group(
            value
        )

    return groups


# ============================================================
# Merged groups
# ============================================================

def _get_group_columns_from_merged_region(
    sheet,
    row_index: int,
    col_index: int,
    groups: dict[int, str],
) -> list[int]:

    merged = get_merged_region(
        sheet,
        row_index,
        col_index,
    )

    if not merged:

        return []

    (
        row_start,
        row_end,
        col_start,
        col_end,
    ) = merged

    # Объединение должно находиться
    # в одной строке.
    if (
        row_end - row_start
        != 1
    ):

        return []

    result = []

    for group_col in groups:

        if (
            col_start
            <= group_col
            < col_end
        ):

            result.append(
                group_col
            )

    return result


# ============================================================
# Lesson
# ============================================================

def create_lesson_record(
    group_name,
    current_day,
    current_date,
    lesson_number,
    lesson_time,
    sheet_name,
    lesson_text,
    lesson_info,
    lesson_type,
    schedule_name,
    schedule_type,
    schedule_key,
):

    return {
        "group": group_name,
        "day": current_day,
        "date": current_date,
        "lesson_number": lesson_number,
        "lesson_time": lesson_time,
        "sheet": sheet_name,

        "raw_text": lesson_text,

        "subject": lesson_info[
            "subject"
        ],

        "teacher": lesson_info[
            "teacher"
        ],

        "room": lesson_info[
            "room"
        ],

        "building": lesson_info[
            "building"
        ],

        "is_online": lesson_info[
            "is_online"
        ],

        "lesson_type": lesson_type,

        "schedule_name": schedule_name,
        "schedule_type": schedule_type,
        "schedule_key": schedule_key,
    }


# ============================================================
# Main parser
# ============================================================

def parse_excel(
    file_path: str,
    schedule_name: str,
    schedule_type: str,
    schedule_key: str,
) -> list[dict]:

    workbook = _open_workbook(
        file_path
    )

    parsed_lessons: list[dict] = []

    for sheet in _get_sheets(
        workbook
    ):

        sheet_name = _get_sheet_name(
            sheet
        )

        print(
            f"Processing sheet: {sheet_name}"
        )

        groups = _get_groups(
            sheet
        )

        current_day = None
        current_date = None

        processed_merged_cells = set()

        for row_index in range(
            3,
            _get_nrows(sheet),
        ):

            day_cell = get_cell_value(
                sheet,
                row_index,
                0,
            )

            time_cell = get_cell_value(
                sheet,
                row_index,
                1,
            )

            if day_cell:

                (
                    current_day,
                    current_date,
                ) = parse_day(
                    str(day_cell)
                )

            if not time_cell:

                continue

            (
                lesson_number,
                lesson_time,
            ) = parse_time(
                str(time_cell)
            )

            for col_index, group_name in groups.items():

                merged_region = get_merged_region(
                    sheet,
                    row_index,
                    col_index,
                )

                if merged_region:

                    (
                        row_start,
                        row_end,
                        col_start,
                        col_end,
                    ) = merged_region

                    # Для горизонтального merge нас интересует
                    # только верхняя строка.
                    if (
                        row_end - row_start
                        != 1
                    ):

                        continue

                    # Только первая группа данного merge
                    # обрабатывает всю область.
                    if col_index != col_start:

                        continue

                    merged_key = (
                        row_index,
                        col_start,
                        col_end,
                    )

                    if (
                        merged_key
                        in processed_merged_cells
                    ):

                        continue

                    processed_merged_cells.add(
                        merged_key
                    )

                    lesson_cell = get_cell_value(
                        sheet,
                        row_index,
                        col_index,
                    )

                    if not lesson_cell:

                        continue

                    lesson_text = normalize_whitespace(
                        str(lesson_cell)
                    )

                    if not lesson_text:

                        continue

                    is_lecture = is_underlined(
                        sheet,
                        row_index,
                        col_index,
                    )

                    lesson_type = (
                        "Лекция"
                        if is_lecture
                        else "Семинар"
                    )

                    # shared нужен только для lesson_parser
                    lesson_info = parse_lesson_text(
                        lesson_text,
                        is_shared=True,
                    )

                    if lesson_info["skip"]:

                        continue

                    merged_group_columns = (
                        _get_group_columns_from_merged_region(
                            sheet,
                            row_index,
                            col_index,
                            groups,
                        )
                    )

                    if not merged_group_columns:

                        merged_group_columns = [
                            col_index
                        ]

                    for shared_col in merged_group_columns:

                        shared_group = groups[
                            shared_col
                        ]

                        parsed_lessons.append(
                            create_lesson_record(
                                shared_group,
                                current_day,
                                current_date,
                                lesson_number,
                                lesson_time,
                                sheet_name,
                                lesson_text,
                                lesson_info,
                                lesson_type,
                                schedule_name,
                                schedule_type,
                                schedule_key,
                            )
                        )

                    continue

                # ------------------------------------------------
                # Обычная ячейка
                # ------------------------------------------------

                lesson_cell = get_cell_value(
                    sheet,
                    row_index,
                    col_index,
                )

                if not lesson_cell:

                    continue

                lesson_text = normalize_whitespace(
                    str(lesson_cell)
                )

                if not lesson_text:

                    continue

                # --------------------------------------------
                # Для обычной пары тип определяется
                # непосредственно по её форматированию.
                # --------------------------------------------

                is_lecture = is_underlined(
                    sheet,
                    row_index,
                    col_index,
                )

                lesson_type = (
                    "Лекция"
                    if is_lecture
                    else "Семинар"
                )

                lesson_info = parse_lesson_text(
                    lesson_text,
                    is_shared=False,
                )

                if lesson_info["skip"]:

                    continue

                parsed_lessons.append(
                    create_lesson_record(
                        group_name,
                        current_day,
                        current_date,
                        lesson_number,
                        lesson_time,
                        sheet_name,
                        lesson_text,
                        lesson_info,
                        lesson_type,
                        schedule_name,
                        schedule_type,
                        schedule_key,
                    )
                )

    return parsed_lessons


def export_to_json(
    lessons: list[dict],
    output_path: str,
) -> None:

    with open(
        output_path,
        "w",
        encoding="utf-8",
    ) as file:

        json.dump(
            lessons,
            file,
            ensure_ascii=False,
            indent=4,
        )
